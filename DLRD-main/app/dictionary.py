"""Deterministic variable-dictionary ingestion — business labels onto nodes.

Some projects ship a tabular *data dictionary* (CSV / Excel) that maps technical
names (a column, a table) to a human/business label and a description. This
module reads those auxiliary files and attaches their rows as ATTRIBUTES on the
graph nodes whose name matches, so the dashboard, story and chat can surface the
business meaning of an otherwise cryptic identifier.

Hard guarantees (kept by construction):
  * Deterministic. No LLM, no network, no new nodes, no new edges — only
    string attributes added onto already-built table/column nodes.
  * Parse-only. CSV via the stdlib ``csv`` module; ``.xlsx``/``.xlsm`` via
    ``openpyxl`` reading VALUES ONLY (``data_only=True``) — formulas are never
    evaluated and macros are never executed (openpyxl cannot run them).
  * Schema-agnostic. No column name is hard-coded: the technical-key column is
    auto-detected by how many of its values match existing node names. A file
    where no column clears the match threshold is treated as not-a-dictionary
    and skipped, so arbitrary data files cannot pollute the graph.
  * Bounded + safe. Files must live inside the project root (symlinked
    directories are not followed); file count, size, rows, columns, attribute
    count and value length are all capped. Any failure skips the file (counted),
    never raises.

The returned stats carry counts only (no names, no paths, no values), so they
are safe to log next to the rest of the run diagnostics.
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
from pathlib import Path

from .scanner import IGNORE_DIRS

log = logging.getLogger("data_lineage_retro_documentation.dictionary")

# Auto-detection: a column is accepted as the technical key when at least this
# fraction of its non-empty values match an existing table/column node name,
# AND it produces at least this many distinct matches (so a 1-row coincidence
# never qualifies a random data file as a dictionary). The rate is deliberately
# low: an enterprise/global dictionary may cover thousands of variables of which
# only a small fraction appear in any one project, so a high bar would silently
# reject a legitimate dictionary. The "distinct matches" guard is what keeps a
# random data file (whose values match no node at all) from being accepted.
_MATCH_RATE_THRESHOLD = 0.1
_MIN_DISTINCT_MATCHES = 2

# Bounds — keep ingestion cheap and prevent a pathological file from ballooning
# the graph. None of these are configurable from analyzed content.
_MAX_FILES = 50            # candidate dictionary files scanned per project
_MAX_FILE_BYTES = 8_000_000
_MAX_TABLES_PER_FILE = 12  # worksheets in a workbook
_MAX_ROWS = 20_000
_MAX_COLS = 64
_MAX_ATTRS_PER_NODE = 12
_MAX_KEY_LEN = 40
_MAX_VALUE_LEN = 200

_CANDIDATE_EXTS = (".csv", ".xlsx", ".xlsm")
_VCS_DIRS = {".git", ".hg", ".svn", ".bzr"}
_SKIP_DIRS = IGNORE_DIRS | _VCS_DIRS
_CONFIG_NAMES = ("dlrd_dictionary.json", ".dlrd_dictionary.json")

_KEY_SANITIZE_RE = re.compile(r"[^a-z0-9]+")


def _new_stats() -> dict:
    return {
        "files_seen": 0,
        "dictionaries_detected": 0,
        "files_skipped": 0,
        "nodes_enriched": 0,
        "attributes_written": 0,
    }


def _sanitize_key(header: str) -> str:
    """Derive a safe, stable attribute key from a column header.

    Lowercased, non-alphanumerics collapsed to single underscores, trimmed to a
    sane length. Returns "" when nothing usable remains (caller skips it)."""
    key = _KEY_SANITIZE_RE.sub("_", header.strip().lower()).strip("_")
    return key[:_MAX_KEY_LEN]


def _clean_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > _MAX_VALUE_LEN:
        text = text[:_MAX_VALUE_LEN].rstrip() + "…"
    return text


def _build_name_index(graph: dict) -> dict[str, list[dict]]:
    """Map a matchable name -> the table/column node dicts carrying it.

    Indexes column names and table names; also indexes the last dotted segment
    of a qualified table name (``demo.t_orders`` -> ``t_orders``) so a
    dictionary that lists the bare table name still matches."""
    index: dict[str, list[dict]] = {}

    def add(name: str, node: dict) -> None:
        name = (name or "").strip()
        if name:
            index.setdefault(name, []).append(node)

    for node in graph.get("nodes", []):
        ntype = node.get("type")
        if ntype not in ("table", "column"):
            continue
        name = node.get("name", "")
        add(name, node)
        if ntype == "table" and "." in name:
            add(name.rsplit(".", 1)[-1], node)
    return index


# --------------------------------------------------------------------------- IO

def _iter_candidate_files(root: Path):
    """Yield candidate dictionary file paths within ``root`` (symlinked dirs not
    followed, ignored/VCS dirs pruned, capped count + size)."""
    yielded = 0
    for dirpath, dirnames, filenames in os.walk(root, followlinks=False):
        # Prune ignored / version-control / symlinked directories in place.
        dirnames[:] = [
            d for d in dirnames
            if d not in _SKIP_DIRS
            and not d.startswith(".")  # hidden dirs (kept simple + safe)
            and not os.path.islink(os.path.join(dirpath, d))
        ]
        for fn in filenames:
            if not fn.lower().endswith(_CANDIDATE_EXTS):
                continue
            full = Path(dirpath) / fn
            try:
                if full.is_symlink() or not full.is_file():
                    continue
                if full.stat().st_size > _MAX_FILE_BYTES:
                    continue
            except OSError:
                continue
            yield full
            yielded += 1
            if yielded >= _MAX_FILES:
                return


def _read_csv_table(path: Path) -> list[tuple[list[str], list[list[str]]]]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        reader = csv.reader(fh)
        rows = []
        for i, row in enumerate(reader):
            if i > _MAX_ROWS:
                break
            rows.append([_clean_value(c) for c in row[:_MAX_COLS]])
    if not rows:
        return []
    header, body = rows[0], rows[1:]
    return [(header, body)]


def _read_excel_tables(path: Path) -> list[tuple[list[str], list[list[str]]]]:
    """Read each worksheet as a (header, rows) table — VALUES ONLY.

    ``data_only=True`` returns the last cached value for formula cells (openpyxl
    never computes them); ``read_only=True`` streams rows. Macros are never run.
    Returns [] when openpyxl is unavailable so CSV ingestion still works."""
    try:
        from openpyxl import load_workbook
    except Exception:  # noqa: BLE001 - optional dependency; degrade to CSV-only
        log.info("openpyxl not installed — Excel dictionaries skipped (CSV still ingested).")
        return []

    tables: list[tuple[list[str], list[list[str]]]] = []
    wb = None
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True, keep_vba=False)
        for ws in wb.worksheets[:_MAX_TABLES_PER_FILE]:
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > _MAX_ROWS:
                    break
                rows.append([_clean_value(c) for c in list(row)[:_MAX_COLS]])
            if rows:
                tables.append((rows[0], rows[1:]))
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass
    return tables


def _read_tables(path: Path) -> list[tuple[list[str], list[list[str]]]]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _read_csv_table(path)
    if ext in (".xlsx", ".xlsm"):
        return _read_excel_tables(path)
    return []


# ----------------------------------------------------------------- config (opt)

def _load_config(root: Path) -> dict[str, dict]:
    """Optional per-project config: posix-relative-path -> {key, labels}.

    Tolerant: any missing file / parse error yields an empty map (auto-detect).
    Accepts either {"files": [{path,keyColumn,labelColumns}]} or a single
    {file,keyColumn,labelColumns} object."""
    for name in _CONFIG_NAMES:
        cfg_path = root / name
        if not cfg_path.is_file():
            continue
        try:
            data = json.loads(cfg_path.read_text(encoding="utf-8", errors="replace"))
        except Exception:  # noqa: BLE001
            return {}
        entries = data.get("files") if isinstance(data, dict) else None
        if entries is None and isinstance(data, dict) and data.get("file"):
            entries = [data]
        out: dict[str, dict] = {}
        for e in entries or []:
            if not isinstance(e, dict):
                continue
            rel = str(e.get("path") or e.get("file") or "").replace("\\", "/").strip().lstrip("./")
            if not rel:
                continue
            labels = e.get("labelColumns") or e.get("labels") or []
            out[rel.lower()] = {
                "key": str(e.get("keyColumn") or e.get("key") or "").strip(),
                "labels": [str(c).strip() for c in labels if str(c).strip()],
            }
        return out
    return {}


# -------------------------------------------------------------- key detection

def _column_values(body: list[list[str]], col: int) -> list[str]:
    out = []
    for row in body:
        if col < len(row):
            v = row[col].strip()
            if v:
                out.append(v)
    return out


def _detect_key_column(
    header: list[str], body: list[list[str]], name_index: dict[str, list[dict]],
) -> int:
    """Index of the column whose values best match node names, or -1 if none
    clears the threshold."""
    best_col, best_rate = -1, 0.0
    for col in range(len(header)):
        values = _column_values(body, col)
        if not values:
            continue
        distinct_matches = {v for v in values if v in name_index}
        rate = sum(1 for v in values if v in name_index) / len(values)
        if (
            rate >= _MATCH_RATE_THRESHOLD
            and len(distinct_matches) >= _MIN_DISTINCT_MATCHES
            and rate > best_rate
        ):
            best_col, best_rate = col, rate
    return best_col


def _resolve_key_column(
    header: list[str], body: list[list[str]], name_index: dict[str, list[dict]],
    cfg: dict | None,
) -> tuple[int, set[int]]:
    """Return (key_column_index, restricted_label_columns).

    A config entry naming an existing key column wins; otherwise auto-detect.
    ``restricted_label_columns`` is empty unless the config restricts which
    other columns become attributes."""
    header_lower = [h.strip().lower() for h in header]
    if cfg and cfg.get("key"):
        try:
            key_idx = header_lower.index(cfg["key"].lower())
        except ValueError:
            key_idx = -1
        if key_idx >= 0:
            label_cols: set[int] = set()
            for lbl in cfg.get("labels", []):
                if lbl.lower() in header_lower:
                    label_cols.add(header_lower.index(lbl.lower()))
            return key_idx, label_cols
    return _detect_key_column(header, body, name_index), set()


# --------------------------------------------------------------------- attach

def _attach_table(
    header: list[str], body: list[list[str]], key_idx: int, label_cols: set[int],
    name_index: dict[str, list[dict]], stats: dict, enriched_nodes: set[str],
) -> None:
    # Pre-sanitize the attribute key for every non-key column.
    attr_keys: dict[int, str] = {}
    for col, raw_header in enumerate(header):
        if col == key_idx:
            continue
        if label_cols and col not in label_cols:
            continue
        key = _sanitize_key(raw_header) or f"attr_{col}"
        attr_keys[col] = key

    for row in body:
        if key_idx >= len(row):
            continue
        tech_name = row[key_idx].strip()
        if not tech_name:
            continue
        targets = name_index.get(tech_name)
        if not targets:
            continue
        for col, attr_key in attr_keys.items():
            if col >= len(row):
                continue
            value = row[col].strip()
            if not value:
                continue
            for node in targets:
                attrs = node.setdefault("attributes", {})
                if attr_key in attrs:
                    continue  # first dictionary wins; don't double-count
                if len(attrs) >= _MAX_ATTRS_PER_NODE:
                    continue
                attrs[attr_key] = value
                stats["attributes_written"] += 1
                enriched_nodes.add(node["id"])


# ------------------------------------------------------------------- public API

def enrich_nodes_with_dictionaries(graph: dict, project_root: str) -> dict:
    """Attach business-label attributes from project dictionary files onto
    matching table/column nodes. Mutates ``graph`` in place; returns count-only
    stats. Never raises."""
    stats = _new_stats()
    try:
        root = Path(project_root).resolve()
        if not root.is_dir():
            return stats
        name_index = _build_name_index(graph)
        if not name_index:
            return stats  # no table/column nodes -> nothing to match against

        config = _load_config(root)
        enriched_nodes: set[str] = set()

        for path in _iter_candidate_files(root):
            stats["files_seen"] += 1
            try:
                rel = path.resolve().relative_to(root).as_posix().lower()
            except (ValueError, OSError):
                stats["files_skipped"] += 1
                continue
            cfg = config.get(rel)
            try:
                tables = _read_tables(path)
            except Exception as exc:  # noqa: BLE001 - unreadable file -> skip
                log.debug("dictionary read failed: %s", exc)
                stats["files_skipped"] += 1
                continue

            detected_here = False
            for header, body in tables[:_MAX_TABLES_PER_FILE]:
                if not header or not body:
                    continue
                key_idx, label_cols = _resolve_key_column(header, body, name_index, cfg)
                if key_idx < 0:
                    continue  # not a dictionary table (or config key absent)
                detected_here = True
                _attach_table(header, body, key_idx, label_cols, name_index, stats, enriched_nodes)

            if detected_here:
                stats["dictionaries_detected"] += 1
            else:
                stats["files_skipped"] += 1

        stats["nodes_enriched"] = len(enriched_nodes)
    except Exception as exc:  # noqa: BLE001 - ingestion is best-effort, never fatal
        log.debug("dictionary enrichment failed: %s", exc)
    return stats
